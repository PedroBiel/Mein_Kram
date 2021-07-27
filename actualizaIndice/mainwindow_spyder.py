# -*- coding: utf-8 -*-
"""
Actualiza la hoja Excel 'Índice de archivos.xlsx'

Created on Tue Mar  9 10:47:22 2021

__author__ = Pedro Biel
__version__ = 0.0.0
__email__ = pbiel@taimweser.com
"""


# import qdarkstyle
import os
import sys

import time
import xlsxwriter

from itertools import groupby
from openpyxl import load_workbook

from PyQt5 import uic
from PyQt5.QtCore import QLibraryInfo, QLocale, QTranslator
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QApplication, QMainWindow

from controlador.cntactualiza import CntActualiza


class MainWindow(QMainWindow):

    def __init__(self, parent=None):

        QMainWindow.__init__(self, parent)
        uic.loadUi('../resources/base/actualiza_ind.ui', self)
        
        # Entorno virtual.
        print('\nEntorno virtual:', sys.prefix)
        print('python', sys.version)
        
        print(time.strftime('%d.%m.%Y'))
        
        # Librerías.
        self.os = os
        self.groupby = groupby
        self.load_workbook = load_workbook
        self.time = time
        self.xlsxwriter = xlsxwriter
        
        # Parámetros.
        
        # Widgets PyQt5.
        self.lbl_actualiza = self.labelActualiza
        self.btn_actualiza = self.pushButtonActualiza
        self.lbl_mensaje_1 = self.labelMensaje1
        self.lbl_mensaje_2 = self.labelMensaje2
        self.lbl_mensaje_3 = self.labelMensaje3
        self.lbl_mensaje_4 = self.labelMensaje4
        
        # Instancias de clase.
        self.cnt = CntActualiza(self)
        
        # Evento para actualizar el índice.
        self.btn_actualiza.clicked.connect(self.cnt.actualiza)
        
        # Status bar.
        text = 'Tus deseos son ordenes.'
        self.status_bar(text)
        
    # Salidas PyQt5.
    
    def status_bar(self, text):
        """Texto informativo en status bar."""
        
        self.statusbar.setFont(QFont('Consolas', 8))
        self.statusbar.showMessage(text)

    def mensaje_yoda(self):
        """Mensaje del maestro Yoda."""
        
        self.lbl_mensaje_1.setText('El Maestro Yoda dice:')
        self.lbl_mensaje_2.setText(
            'Fichero con el índice de archivos actualizdo.'
            )
        self.lbl_mensaje_3.setText('Nombre: Índice de archivos.xlsx')
        self.lbl_mensaje_4.setText(
            'Ruta: H:\COMUN MT\5.CE\Aplicaciones de cálculo'
            )
        
if __name__ == '__main__':

    app = QApplication(sys.argv)

    qt_translator = QTranslator()
    qt_translator.load(
        'qtbase_' + QLocale.system().name(),
        QLibraryInfo.location(QLibraryInfo.TranslationsPath)
        )
    app.installTranslator(qt_translator)

    # app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())

    myapp = MainWindow()
    myapp.show()
    sys.exit(app.exec_())
